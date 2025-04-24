import streamlit as st
import os
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import time

# Função para ajustar largura das colunas no Excel
def ajustar_largura_colunas_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  
        ws.column_dimensions[column].width = adjusted_width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    wb.save(caminho_arquivo)

# Função para converter CSV em Excel
def csv_excel(input_csv, output_folder):
    try:
        dados_csv = pd.read_csv(input_csv)
        output_excel = os.path.join(output_folder, "ResultadosExcel.xlsx")
        dados_csv.to_excel(output_excel, index=False)
        return output_excel
    except Exception as e:
        st.error(f"Erro ao converter CSV para Excel: {str(e)}")
        return None

# Processamento do Excel
def calcular_conforto_noturno(arquivo_excel):
    df = pd.read_excel(arquivo_excel)
    df = df.drop("Environment:Site Outdoor Air Drybulb Temperature [C](TimeStep)", axis=1)

    # Converter 'Date/Time' para datetime
    df['DateTime'] = pd.to_datetime(
        df['Date/Time'].str.strip(),
        format='%m/%d %H:%M:%S',
        errors='coerce'
    )
    df = df.dropna(subset=['DateTime'])

    # Filtrar período noturno (18h às 6h)
    df['Hora'] = df['DateTime'].dt.hour
    df_noite = df[(df['Hora'] >= 18) | (df['Hora'] < 6)]

    # Selecionar colunas de temperatura
    cols_temp = [col for col in df_noite.columns if any(x in col for x in ['Temp', 'Temperature'])]

    # Verificar se existem colunas de temperatura
    if not cols_temp:
        print("Nenhuma coluna de temperatura encontrada!")
    else:
        print(f"Colunas de temperatura encontradas: {cols_temp}")
        
        # Converter colunas para numérico (ignorando erros)
        for col in cols_temp:
            df_noite[col] = pd.to_numeric(df_noite[col], errors='coerce')
        
        # Filtrar linhas onde PELO MENOS UMA coluna de temperatura está entre 18 e 28
        mask = df_noite[cols_temp].apply(lambda x: (x >= 18) & (x <= 28)).any(axis=1)
        indices = df_noite.index[mask].tolist()
        
        df_noite = df_noite.loc[indices, cols_temp + ['DateTime']]

        horas_conforto = df_noite.iloc[:, 0].count() * 0.25
        return horas_conforto



# Configuração do Streamlit
st.title("🏠 EnergyPlus Simulation via Streamlit")
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# Upload de arquivos e seleção de arquivos existentes
existing_files = os.listdir(UPLOAD_FOLDER)
epw_file = st.selectbox("Escolha um arquivo pré-existente na plataforma:", ["Nenhum arquivo selecionado"] + existing_files)


if epw_file != "Nenhum arquivo selecionado":
    file_path = os.path.join(UPLOAD_FOLDER, epw_file)
    epw_path = os.path.join(UPLOAD_FOLDER, "weather.epw")
else:
    st.write(f"Você não selecionou nenhum arquivo epw (selecione algum ou faça upload do mesmo): ")
    epw_file = st.file_uploader("Envie o arquivo .epw", type=["epw"])
    if epw_file:
        epw_path = os.path.join(UPLOAD_FOLDER, "weather.epw")
        with open(epw_path, "wb") as f:
            f.write(epw_file.getbuffer())
    
    

idf_file = st.file_uploader("Envie o arquivo .idf", type=["idf"])

if idf_file and epw_file:
    st.success("Arquivos carregados com sucesso!")
    idf_path = os.path.join(UPLOAD_FOLDER, "input.idf")
    with open(idf_path, "wb") as f:
        f.write(idf_file.getbuffer())
    
    
    if st.button("🔄 Rodar Simulação"):
        with st.spinner("Executando EnergyPlus..."):
            try:
                command = ["EnergyPlus", "-r", "-d", OUTPUT_FOLDER, "-w", epw_path, "--expandobjects", "--readvars", idf_path]
                result = subprocess.run(command, capture_output=True, text=True)
                if result.returncode != 0:
                    st.error(f"Erro ao rodar EnergyPlus: {result.stderr}")
            except Exception as e:
                st.error(f"Erro ao executar EnergyPlus: {str(e)}")
            output_csv = os.path.join(OUTPUT_FOLDER, "eplusout.csv")
            if os.path.exists(output_csv):
                output_excel = csv_excel(output_csv, OUTPUT_FOLDER)
                ajustar_largura_colunas_excel(output_excel)

                horas_conforto = calcular_conforto_noturno(output_excel)
                st.write(f"Horas de Conforto Térmico (Noite): {horas_conforto}")

                with open(output_excel, "rb") as f:
                    st.download_button("📥 Baixar Resultados (Excel)", data=f, file_name="resultados.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.error("Arquivo de resultados não encontrado.")
