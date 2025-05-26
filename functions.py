# Importando as bibliotecas necessárias
import streamlit as st  # Interface web
import os  # Manipulação de arquivos e diretórios
import pandas as pd  # Manipulação de dados
import calendar  # Trabalhar com meses
from datetime import datetime, timedelta  # Trabalhar com datas e tempos
from openpyxl import load_workbook  # Manipulação de arquivos Excel
from openpyxl.styles import Alignment  # Estilo de células no Excel

# Função que ajusta automaticamente a largura das colunas do Excel
def ajustar_largura_colunas_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    # Ajustar largura com base no conteúdo
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    # Centralizar o texto das células
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    wb.save(caminho_arquivo)

# Função para converter um arquivo CSV em Excel
def csv_excel(input_csv, output_folder):
    try:
        dados_csv = pd.read_csv(input_csv)
        output_excel = os.path.join(output_folder, "ResultadosExcel.xlsx")
        dados_csv.to_excel(output_excel, index=False)
        return output_excel
    except Exception as e:
        st.error(f"Erro ao converter CSV para Excel: {str(e)}")
        return None

# Função para inverter datas no formato MM/DD ou MM/DD/YYYY para DD/MM ou DD/MM/YYYY
def inverter_data(data_str):
    partes = data_str.split('/')
    if len(partes) == 2:
        mes, dia = partes
        return f"{dia}/{mes}"
    elif len(partes) == 3:
        mes, dia, ano = partes
        return f"{dia}/{mes}/{ano}"
    return data_str

# Ajusta colunas 'Data' e 'Hora' a partir da coluna 'Date/Time'
def formatar_data_hora(df, ano=2025):
    date = []
    hour = []
    for value in df['Date/Time'].astype(str):
        parts = value.strip().split()
        if len(parts) == 2:
            date_str, hour_str = parts
        else:
            date_str = parts[0]
            hour_str = '00:00:00'
        data_formatada = inverter_data(date_str)
        if len(data_formatada.split('/')) == 2:
            data_formatada = f"{data_formatada}/{ano}"
        date.append(data_formatada)
        hour.append(hour_str)
    df.drop(columns=['Date/Time'], inplace=True)
    df.insert(0, "Hora", hour)
    df.insert(0, "Data", date)
    return df

# Converte 'Data' e 'Hora' para um datetime real
def formatar_para_datetime(df):
    datas_reais = []
    for _, row in df.iterrows():
        try:
            if pd.isna(row['Data']) or pd.isna(row['Hora']):
                raise ValueError("Valores ausentes.")
            date_str = row['Data']
            time_str = row['Hora']
            partes_data = date_str.split('/')
            if len(partes_data) != 3:
                raise ValueError("Formato inválido.")
            day, month, year = map(int, partes_data)
            base_date = datetime(year=year, month=month, day=day)
            if time_str.startswith('24:'):
                base_date += timedelta(days=1)
                time_str = '00' + time_str[2:]
            hora, minuto, segundo = map(int, time_str.split(':'))
            data_certa = datetime(base_date.year, base_date.month, base_date.day, hora, minuto, segundo)
            datas_reais.append(pd.Timestamp(data_certa))
        except:
            datas_reais.append(pd.NaT)
    return datas_reais

# Processa colunas temporais e cria coluna 'DateTime'
def processar_dados_temporais(df, cols_temp):
    for _, col in cols_temp:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df['DateTime'] = formatar_para_datetime(df)
    df = df[df['DateTime'].notna()].copy()
    if df['DateTime'].empty:
        print("Nenhuma data válida encontrada!")
        return df
    df['Month'] = df['DateTime'].dt.month
    df['Month_Name'] = df['DateTime'].dt.month_name()
    df['Hour'] = df['DateTime'].dt.hour
    df['Day'] = df['DateTime'].dt.day
    return df

# Calcula métricas de conforto térmico
def calcular_metricas_conforto(data, temp_col, CONFORT_MIN=18, CONFORT_MAX=29, INTERVALO_MINUTOS=15):
    HORAS_POR_INTERVALO = INTERVALO_MINUTOS / 60
    if data.empty:
        return {k: 0 for k in [
            'Total de Horas', 'Conforto (Dia)', 'Conforto (Noite)',
            'Total Conforto', 'Sem Conforto', '% Conforto', '% Sem Conforto']}
    total_horas = len(data) * HORAS_POR_INTERVALO
    periodo_dia = (data['Hour'] >= 6) & (data['Hour'] < 18)
    in_conforto = data[temp_col].between(CONFORT_MIN, CONFORT_MAX)
    conforto_dia = (periodo_dia & in_conforto).sum() * HORAS_POR_INTERVALO
    conforto_noite = ((~periodo_dia) & in_conforto).sum() * HORAS_POR_INTERVALO
    total_conforto = conforto_dia + conforto_noite
    return {
        'Total de Horas': round(total_horas, 2),
        'Conforto (Dia)': round(conforto_dia, 2),
        'Conforto (Noite)': round(conforto_noite, 2),
        'Total Conforto': round(total_conforto, 2),
        'Sem Conforto': round(total_horas - total_conforto, 2),
        '% Conforto': round((total_conforto / total_horas) * 100, 1) if total_horas else 0,
        '% Sem Conforto': round(100 - ((total_conforto / total_horas) * 100), 1) if total_horas else 0
    }

# Processa arquivo de temperatura, gera relatório de conforto
def processar_arquivo_temperatura(df, ano=2025):
    df = formatar_data_hora(df, ano)
    cols_temp = [(col.split(':')[0], col) for col in df.columns if "Mean Air Temperature" in col]
    df = processar_dados_temporais(df, cols_temp)
    if df.empty:
        print("Nenhum dado processado para gerar relatório.")
        return pd.DataFrame()
    resultados = []
    for mes_num in range(1, 13):
        mes_nome = calendar.month_name[mes_num]
        dados_mes = df[df['Month'] == mes_num]
        for local, temp_col in cols_temp:
            if dados_mes.empty or temp_col not in dados_mes.columns:
                metricas = {k: 0 for k in [
                    'Total de Horas', 'Conforto (Dia)', 'Conforto (Noite)',
                    'Total Conforto', 'Sem Conforto', '% Conforto', '% Sem Conforto']}
            else:
                metricas = calcular_metricas_conforto(dados_mes, temp_col)
            resultados.append({'Mês': mes_nome, 'Local': local, **metricas})
    relatorio = pd.DataFrame(resultados)
    return relatorio

# Extrai valores entre marcadores específicos de um arquivo texto
def get_values(arquivo, inicio, final):
    capturar = False
    linhas_finais = []
    with open(arquivo, "r", encoding="utf-8") as data:
        linhas = data.readlines()
        for linha_ in linhas:
            if inicio in linha_:
                capturar = True
                continue
            if final in linha_:
                capturar = False
                continue
            if capturar:
                linhas_finais.append(linha_)
        lista_info = [info.split(",")[0] for info in linhas_finais]
        return lista_info

# Adiciona espaçamento e formatação a uma lista
def add_espaco(lista, espaco, espaco2):
    lista = [str(item).rjust(len(str(item)) + espaco) for item in lista]
    lista = [str(item) + ',' for item in lista]
    lista = [str(item).ljust(16 + espaco2) for item in lista]
    return lista

# Mescla duas listas combinando valores e rótulos
def mesclar_listas(list, list__):
    lista_mesclada = []
    for sublista in list__:
        mesclado = [f"{valor} {rotulo}" for valor, rotulo in zip(sublista[2:], list[1:])]
        lista_mesclada.append(mesclado)
    return lista_mesclada

# Exclui itens específicos entre marcadores de um arquivo texto
def excluir_material(arquivo, inicio, final):
    capturar = False
    linhas_finais = []
    with open(arquivo, "r", encoding="utf-8") as data:
        linhas = data.readlines()
        for linha_ in linhas:
            if inicio in linha_:
                capturar = True
                linhas_finais.append(linha_)
                continue
            if final in linha_:
                capturar = False
                linhas_finais.append(linha_)
                continue
            if not capturar:
                linhas_finais.append(linha_)
        return linhas_finais

# Adiciona novos elementos entre marcadores em um arquivo texto
def add_novos_elementos(arquivo, lista_mesclada, inicio, final):
    with open(arquivo, "r", encoding="utf-8") as data:
        linhas = data.readlines()
    nasecao = False
    novo_arquivo = []
    for linha in linhas:
        novo_arquivo.append(linha)
        if inicio in linha:
            nasecao = True
        if final in linha:
            nasecao = False
            continue
        if nasecao:
            for lista in lista_mesclada:
                for elemento in lista:
                    novo_arquivo.append(f"{elemento}\n")
    with open(arquivo, "w", encoding="utf-8") as data:
        data.writelines(novo_arquivo)
